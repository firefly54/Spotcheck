#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
lampclicked = 0
viewresultclicked = 0
temp_label = 0
name = "/"
entry_num = 0
wait = 0
pos_result = list(range(48))
path0 = "/"
path1 = "/"
path2 = "/"
path3 = "/"
path4 = "/"
path5 = "/"
foldername = ""
importfilename = ""
id_list = list(range(48))
samples = 0
lamp_createclicked = 0
lampdir_old = ""
div = list(range(48))
start_point = (0,0)
end_point = (0,0)
thr_set= 15
#thr2_set = 15
setid48clicked = 0
setid25clicked = 0
########################################################### GLOBAL VARIABLE - END ##################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')
########################################################### MAIN WINDOW INIT - END #################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)
############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, start_point=(283,85), end_point=(497,372)):
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))

    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][0]+list_bgrvalue[i][j][1]+list_bgrvalue[i][j][2]/3)))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    global thr_set, id_list
    for i in range(len(sorted_contours1)):
        if(id_list[i]=='N/A'):
            cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
        else:
            if(result_list[i]<=10):
                cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
            else:
                if(result_list[i] <= float(thr_set)):
                    cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
                else:
                    cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)

    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    try:
        subprocess.Popen(['killall','florence'])
    except:
        pass
    root.attributes('-fullscreen', True)

    buttonFont = font.Font(family='Helvetica', size=10, weight='bold')
    global mainscreen_labelframe
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)

    logo_img = Image.open('/home/pi/Spotcheck/logo.png')
    logo_width, logo_height = logo_img.size
    scale_percent = 50
    width = int(logo_width * scale_percent / 100)
    height = int(logo_height * scale_percent / 100)
    display_img = logo_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
    logo_label.image = image_select
    logo_label.place(x=5,y=5)

    sc_label = Label(mainscreen_labelframe, font=("Courier",45,'bold'), fg='dodger blue', bg='white',text='SPOTCHECK')
    sc_label.place(x=227, y=65)
    lampversion_label = Label(mainscreen_labelframe, font=("Courier",13,'bold'), fg='grey88', bg='white',text='LAMP VERSION')
    lampversion_label.place(x=431, y=118)

    power_labelframe = LabelFrame(mainscreen_labelframe, bg='white',text='POWER', width=405, height=120)
    power_labelframe.place(x=190,y=350)
    def shutdown_click():
        os.system("sudo shutdown -h now")
    def restart_click():
        os.system("sudo shutdown -r now")
    def exit_click():
        root.destroy()
    exit_button = Button(power_labelframe, fg='white', activebackground="dodger blue", font=('Courier','13','bold'), bg="blue4", text="EXIT", height=3, width=9, borderwidth=0, command=exit_click)
    exit_button.place(x=12,y=8)
    shutdown_button = Button(power_labelframe, fg='white', activebackground="red", font=('Courier','13','bold'), bg="red3", text="SHUTDOWN", height=3, width=9, borderwidth=0, command=shutdown_click)
    shutdown_button.place(x=142,y=8)
    restart_button = Button(power_labelframe, fg='white', activebackground="lawn green", font=('Courier','13','bold'), bg="green", text="RESTART", height=3, width=9, borderwidth=0, command=restart_click)
    restart_button.place(x=272,y=8)

    def newprogram_click():
        setfoldername()

    def setid_click():
        setwells()
        
    newprogram_button = Button(mainscreen_labelframe, bg="grey88", activebackground="grey95", text="NEW PROGRAM", font=buttonFont, borderwidth=0, height=4, width=20,command=newprogram_click)
    newprogram_button.place(x=210,y=220)
    setid_button = Button(mainscreen_labelframe, bg="grey88", activebackground="grey95", text="SET ID", font=buttonFont, borderwidth=0, height=4, width=20,command=setid_click)
    setid_button.place(x=410,y=220)

######################################################## FOLDER NAME SCREEN - START ####################################################################
def setfoldername():
    setfoldername_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    setfoldername_labelframe.place(x=0,y=0)

    global lampclicked
    lampclicked = 1
    enterframe_labelframe = LabelFrame(setfoldername_labelframe, bg='white', width=768, height=175)
    enterframe_labelframe.place(x=14,y=5)
    foldername_label = Label(enterframe_labelframe, bg='white',text='Folder name', fg='black', font=("Courier",14,'bold'))
    foldername_label.place(x=82,y=25)

    global lamp_createclicked
    global lampdir_old
    directory = strftime("LAMP %y-%m-%d %H.%M.%S")
    if(lamp_createclicked == 0):
        directory_label = Label(enterframe_labelframe, bg='grey94',text=directory)
        directory_label.place(x=304,y=2)
        lampdir_old = directory
    else:
        directory_label = Label(enterframe_labelframe, bg='grey94',text=lampdir_old)
        directory_label.place(x=304,y=2)

    def enter_entry(event):
        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', False)
        subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
    global foldername
    foldername_entry = Entry(enterframe_labelframe, font=("Courier",15), width=36)
    foldername_entry.insert(0,foldername)
    foldername_entry.bind("<Button-1>", enter_entry)
    foldername_entry.place(x=85,y=50)
    
    file_label = Label(enterframe_labelframe, bg='white', fg='grey70', font=("Courier",11,'bold'))
    file_label.place(x=203,y=104)

    def import_click():
        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)
        file = filedialog.askopenfile(initialdir='/home/pi/Desktop/Spotcheck ID/', mode='r', filetypes=[('xlsx file','*.xlsx')])
        global importfilename
        filename = file.name
        if file is not None:
            a=0
            for i in range(len(filename)):
                if(filename[i]=='/'):
                    a=i+1
            importfilename = filename[a:len(filename)]
            file_label['text']=importfilename
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                id_list[i] = sheet[pos].value

    import_button = Button(enterframe_labelframe, font=("Courier",10,'bold'), bg="lavender", text="Import ID", height=2, width=9, borderwidth=0, command=import_click)
    import_button.place(x=85,y=93)
    
    def create_click():
        global lamp_createclicked
        lamp_createclicked = 1
        global foldername
        foldername = foldername_entry.get()
        name = strftime(foldername)
        global path0
        global lampdir_old
        path0 = os.path.join("/home/pi/Desktop/Spotcheck Results", lampdir_old +" "+ name)

        if(foldername_entry.get()==""):
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', True)
            msgbox = messagebox.showwarning(" ","Please enter the folder name!" )
        else:
            if(file_label['text']==""):
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', True)
                msgbox = messagebox.showwarning(" ","Please import ID file!" )
            else:
                if os.path.exists(path0):
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    msg = messagebox.askquestion("The folder already exixts", "Do you want to overwrite it?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"Original image")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"Processed image")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"Result Table")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"Sample image")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Temperature program")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        settemp()
                else:
                    try:
                        subprocess.Popen(['killall','florence'])
                    except:
                        pass
                    root.attributes('-fullscreen', True)
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"Original image")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"Processed image")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"Result Table")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"Sample image")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Temperature program")
                    os.mkdir(path5)
                    mainscreen_labelframe.place_forget()
                    settemp()
    create_button = Button(enterframe_labelframe, font=("Courier",16,'bold'), bg="lawn green", text="OK", height=3, width=6, borderwidth=0, command=create_click)
    create_button.place(x=565,y=50)

    def back_click():
        mainscreen()

    back_button = Button(setfoldername_labelframe, font=("Courier",12,'bold'), bg="lavender", text="BACK", height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=190)

########################################################## FOLDER NAME SCREEN - END ################################################################

############################################################# SET WELLS - START ####################################################################
def setwells():
    setwells1_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    setwells1_labelframe.place(x=0,y=0)
    
    setwells2_labelframe = LabelFrame(setwells1_labelframe, bg='white', width=768, height=396)
    setwells2_labelframe.place(x=14,y=5)
    
    setwells3_labelframe = LabelFrame(setwells2_labelframe, text='Number of wells', bg='white', width=120, height=330)
    setwells3_labelframe.place(x=199,y=25)
    
    def button48_click():
        button48['bg'] = 'lawn green'
        button25['bg'] = 'grey94'

        s48_img = Image.open('/home/pi/Spotcheck-Lamp/48well.JPG')
        s48_width, s48_height = s48_img.size
        scale_percent = 76
        width = int(s48_width * scale_percent / 100)
        height = int(s48_height * scale_percent / 100)
        display_img = s48_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        setid_label = Label(setwells2_labelframe, bg='white',image=image_select)
        setid_label.image = image_select
        setid_label.place(x=318,y=30)
    def button25_click():
        button48['bg'] = 'grey94'
        button25['bg'] = 'lawn green'

        s48_img = Image.open('/home/pi/Spotcheck-Lamp/25well.JPG')
        s48_width, s48_height = s48_img.size
        scale_percent = 76
        width = int(s48_width * scale_percent / 100)
        height = int(s48_height * scale_percent / 100)
        display_img = s48_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        setid_label = Label(setwells2_labelframe, bg='white',image=image_select)
        setid_label.image = image_select
        setid_label.place(x=318,y=30)
    
    
    button48 = Button(setwells3_labelframe, activebackground='white', fg='black', font=('Courier','13','bold'), bg="lawn green", text="48", height=4, width=7, borderwidth=0, command=button48_click)
    button48.place(x=9,y=28)
    button25 = Button(setwells3_labelframe, activebackground='white', fg='black', font=('Courier','13','bold'), bg="grey94", text="25", height=4, width=7, borderwidth=0, command=button25_click)
    button25.place(x=9,y=175)
    
    def back_click():
        mainscreen()
    back_button = Button(setwells1_labelframe, font=("Courier",12,'bold'), bg="lavender", text="BACK", height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    
    def ok_click():
        global setid48clicked, setid25clicked
        if(button48['bg']=='lawn green'):
            setid48clicked = 1
            setid25clicked = 0
            setid()

        else:
            setid48clicked = 0
            setid25clicked = 1
            setid()
    ok_button = Button(setwells1_labelframe, fg='black', font=('Courier','12','bold'), bg="lavender", text="OK", height=3, width=11, borderwidth=0, command=ok_click)
    ok_button.place(x=332,y=406)
    button48_click()
############################################################## SET WELLS - END #####################################################################
    
########################################################### SET ID SCREEN - START ##################################################################
def setid():
    global setid48clicked, setid25clicked

    setid1_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    setid1_labelframe.place(x=0,y=0)

    setid2_labelframe = LabelFrame(setid1_labelframe, bg='white', width=470, height=160)
    setid2_labelframe.place(x=320,y=5)

    idpos_label = Label(setid2_labelframe, bg='dodger blue', font=("Courier",24,"bold"))
    idpos_label.place(x=1,y=1)

    setidtable_labelframe = LabelFrame(setid1_labelframe,bg='ghost white', width=600, height=307)
    setidtable_labelframe.place(x=10,y=5)

    def idpos_click(n):
        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        if(idpos_button[n]['bg'] != 'lawn green'):
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                else:
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'white'
        else:
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                if(idpos_button[k]['bg'] == 'grey99'):
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'grey99'

        def enter_entry(event):
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)

        def ok_click(event=None):
            if(id_entry.get()==''):
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', True)
                idpos_button[n]['bg'] = 'lavender'
                if(setid48clicked==0):
                    if(n<=5):
                        idpos_button[n]['text'] = '#'+str(n+1)
                    if(n>5 and n<=11):
                        idpos_button[n]['text'] = '#'+str(n)
                    if(n>11 and n<=17):
                        idpos_button[n]['text'] = '#'+str(n-1)
                    if(n>17 and n<=23):
                        idpos_button[n]['text'] = '#'+str(n-2)
                    if(n>23):
                        idpos_button[n]['text'] = '#'+str(n-3)
                else:
                    idpos_button[n]['text'] = '#'+str(n+1)
                msgbox = messagebox.showwarning(" ","Please enter the ID!")
            else:
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', True)
                idpos_button[n]['text'] = id_entry.get()
                idpos_button[n]['bg'] = 'lawn green'
                try:
                    if(setid48clicked==1):
                        if(n==46):
                            idpos_click(1)
                        else:
                            idpos_click(n+1)
                    else:
                        if(n==4 or n==10 or n==16 or n==22):
                            idpos_click(n+2)
                        elif(n==27):
                            idpos_click(1)
                        else:
                            idpos_click(n+1)

                except:
                    idpos_click(0)

        id_entry = Entry(setid2_labelframe,width=25, font=('Courier',14))
        if(idpos_button[n]['bg'] == 'grey99'):
            id_entry.insert(0,idpos_button[n]['text'])
        id_entry.bind("<Button-1>", enter_entry)
        id_entry.bind("<Return>", ok_click)
        id_entry.place(x=50,y=70)
        id_entry.focus_set()

        setid_label = Label(setid2_labelframe, text='Enter ID', bg='white', font=("Courier",15,"bold"))
        setid_label.place(x=48,y=43)

        if(n<6):
            idpos_label['text'] = 'A' + str(n+1)
        if(n>=6 and n<12):
            idpos_label['text'] = 'B' + str(n-5)
        if(n>=12 and n<18):
            idpos_label['text'] = 'C' + str(n-11)
        if(n>=18 and n<24):
            idpos_label['text'] = 'D' + str(n-17)
        if(n>=24 and n<30):
            idpos_label['text'] = 'E' + str(n-23)
        if(n>=30 and n<36):
            idpos_label['text'] = 'F' + str(n-29)
        if(n>=36 and n<42):
            idpos_label['text'] = 'G' + str(n-35)
        if(n>=42):
            idpos_label['text'] = 'H' + str(n-41)

        ok_button = Button(setid2_labelframe, font=('Courier','12','bold'), bg="lavender", text="OK", height=2, width=8, borderwidth=0, command=ok_click)
        ok_button.place(x=340,y=58)

    idpos_button = list(range(48))
    if(setid48clicked==1):
        h=0
        c=-1
        for i in range(0,48):
            c+=1
            if(i%6==0 and i!=0):
                c=0
                h+=1
            idpos_button[i] = Button(setidtable_labelframe, bg='lavender', justify='left', borderwidth=0, text='#'+str(i+1), width=2, height=2)
            idpos_button[i]['command'] = partial(idpos_click,i)
            idpos_button[i].grid(row=h,column=c,padx=4,pady=4)
            if(i==0):
                idpos_button[i]['state']='disabled'
                idpos_button[i]['bg']= 'red'
            if(i==47):
                idpos_button[i]['state']='disabled'
                idpos_button[i]['bg']= 'green'
    else:
        k=0
        h=0
        c=-1
        for i in range(0,48):
            c+=1
            if(i%6==0 and i!=0):
                c=0
                h+=1
            idpos_button[i] = Button(setidtable_labelframe, bg='lavender', justify='left', borderwidth=0, width=2, height=2)
            idpos_button[i]['command'] = partial(idpos_click,i)
            idpos_button[i].grid(row=h,column=c,padx=4,pady=4)

            if(i==5 or i==11 or i==17 or i==23 or i>=29):
                idpos_button[i]['state']='disabled'
                idpos_button[i]['text']='X'
            elif(i==0):
                k=k+1
                idpos_button[i]['state']='disabled'
                idpos_button[i]['bg']= 'red'
                idpos_button[i]['text']='#'+str(k)
            elif(i==28):
                k=k+1
                idpos_button[i]['state']='disabled'
                idpos_button[i]['bg']= 'green'
                idpos_button[i]['text']='#'+str(k)
            else:
                k=k+1
                idpos_button[i]['text']='#'+str(k)

    def cancel_click():
        msg = messagebox.askquestion("Cancel", "Do you want to cancel without saving?")
        if(msg=="yes"):
            global setid48clicked, setid25clicked
            setid48clicked=0
            setid25clicked=0
            setid1_labelframe.place_forget()
            mainscreen()

    def save_click():
        workbook = Workbook()
        sheet = workbook.active
        for i in range(0,48):
            #pos = "C"+str(i+3)
            if(i<6):
                pos = str(chr(65+i+1))+"2"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5))+"3"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11))+"4"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17))+"5"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23))+"6"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29))+"7"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35))+"8"
            if(i>=42):
                pos = str(chr(65+i-41))+"9"
            if(idpos_button[i]['bg']=='lawn green' or idpos_button[i]['bg']=='grey99'):
                sheet[pos] = idpos_button[i]['text']
            else:
                sheet[pos] = 'N/A'

        global setid48clicked, setid25clicked
        if(setid48clicked==1):
            sheet['B2']='POSC'
            sheet['G9']='NEGC'
        else:
            sheet['B2']='POSC'
            sheet['F6']='NEGC'

        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', False)
        subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)

        if(setid48clicked==1):
            f = filedialog.asksaveasfilename(initialdir='/home/pi/Desktop/Spotcheck ID/48 wells/',defaultextension='.xlsx')
        else:
            f = filedialog.asksaveasfilename(initialdir='/home/pi/Desktop/Spotcheck ID/25 wells/',defaultextension='.xlsx')

        if f is None:
            return
        workbook.save(f)

        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        msg = messagebox.showinfo(' ','Save Done!')
        if(msg=='ok'):
            setid1_labelframe.place_forget()
            mainscreen()

        setid48clicked=0
        setid25clicked=0

    cancel_button = Button(setid1_labelframe, font=('Courier','12','bold'), bg="lavender", text="Cancel" , height=3, width=11, borderwidth=0, command=cancel_click)
    cancel_button.place(x=400,y=170)
    save_button = Button(setid1_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="Save", height=3, width=11, borderwidth=0, command=save_click)
    save_button.place(x=570,y=170)
############################################################ SET ID SCREEN - END ###################################################################

###################################################### SET TEMPERATURES SCREEN - START #############################################################
def settemp():
    fr = open("/home/pi/Spotcheck-Lamp/tempsaved.txt","r")
    t1 = fr.readline()[3:5].strip()
    t2 = fr.readline()[3:5].strip()
    t1m = fr.readline()[4:6].strip()
    t1s = fr.readline()[4:6].strip()
    t2m = fr.readline()[4:6].strip()
    t2s = fr.readline()[4:6].strip()
    thr = fr.readline()[4:8].strip()
    #thr2 = fr.readline()[5:9]

    global samples
    samples=0
    settemp_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    settemp_labelframe.place(x=0,y=0)
    settemptop_labelframe = LabelFrame(settemp_labelframe, bg='white', width=798, height=350)
    settemptop_labelframe.place(x=0,y=52)
    keypad_labelframe = LabelFrame(settemptop_labelframe, bg='white', width=285, height=323)
    keypad_labelframe.place(x=501,y=11)
    title_labelframe = LabelFrame(settemp_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    settemp_label = Label(settemp_labelframe, bg='dodger blue', fg='black', text='SET PARAMETER', font=("Courier",17,'bold'), width=20, height=1 )
    settemp_label.place(x=265,y=12)

    def numpad_click(btn):
        text = "%s" % btn
        if (text!="Delete" and text!="Default"):
            if(entry_num==1):
                t1_entry.insert(END, text)
            if(entry_num==2):
                t2_entry.insert(END, text)
            if(entry_num==3):
                t3_entry.insert(END, text)
            if(entry_num==4):
                t1min_entry.insert(END, text)
            if(entry_num==5):
                t1sec_entry.insert(END, text)
            if(entry_num==6):
                t2min_entry.insert(END, text)
            if(entry_num==7):
                t2sec_entry.insert(END, text)
            if(entry_num==8):
                thr_entry.insert(END, text)

        if text == 'Delete':
            if(entry_num==1):
                t1_entry.delete(0, END)
            if(entry_num==2):
                t2_entry.delete(0, END)
            if(entry_num==3):
                t3_entry.delete(0, END)
            if(entry_num==4):
                t1min_entry.delete(0, END)
            if(entry_num==5):
                t1sec_entry.delete(0, END)
            if(entry_num==6):
                t2min_entry.delete(0, END)
            if(entry_num==7):
                t2sec_entry.delete(0, END)
            if(entry_num==8):
                thr_entry.delete(0, END)

        if text == 'Default':
            if(entry_num==1):
                t1_entry.delete(0, END)
                t1_entry.insert(END, t1)
            if(entry_num==2):
                t2_entry.delete(0, END)
                t2_entry.insert(END, t2)
            if(entry_num==3):
                t3_entry.delete(0, END)
                t3_entry.insert(END, t3)
            if(entry_num==4):
                t1min_entry.delete(0, END)
                t1min_entry.insert(END, t3)
            if(entry_num==5):
                t1sec_entry.delete(0, END)
                t1sec_entry.insert(END, t3)
            if(entry_num==6):
                t2min_entry.delete(0, END)
                t2min_entry.insert(END, t3)
            if(entry_num==7):
                t2sec_entry.delete(0, END)
                t2sec_entry.insert(END, t3)
            if(entry_num==8):
                thr_entry.delete(0, END)
                thr_entry.insert(END, t3)
                
    def numpad():
        global numpad_labelframe
        numpad_labelframe = LabelFrame(keypad_labelframe, bg="white", width=385, height=395)
        numpad_labelframe.place(x=2,y=1)
        button_list = ['7',     '8',      '9',
                       '4',     '5',      '6',
                       '1',     '2',      '3',
                       '0',     '.', 'Delete']
        r = 1
        c = 0
        n = 0
        btn = list(range(len(button_list)))
        for label in button_list:
            cmd = partial(numpad_click, label)
            btn[n] = Button(numpad_labelframe, text=label, font=font.Font(family='Helvetica', size=10, weight='bold'), width=9, height=4, command=cmd)
            btn[n].grid(row=r, column=c, padx=0, pady=0)
            n += 1
            c += 1
            if (c == 3):
                c = 0
                r += 1

    cir_img = Image.open('/home/pi/Spotcheck-Lamp/cir.png')
    cir_width, cir_height = cir_img.size
    scale_percent = 16
    width = int(cir_width * scale_percent / 100)
    height = int(cir_height * scale_percent / 100)
    display_img = cir_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    t1cir_label = Label(settemptop_labelframe, bg='white', image=image_select)
    t1cir_label.image = image_select
    t1cir_label.place(x=70,y=5)
    t2cir_label = Label(settemptop_labelframe, bg='white', image=image_select)
    t2cir_label.image = image_select
    t2cir_label.place(x=275,y=5)

    def entryt1_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 1
        numpad()
    def entryt2_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 2
        numpad()
    def entryt3_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 3
        numpad()
    def entryt1min_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 4
        numpad()
    def entryt1sec_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 5
        numpad()
    def entryt2min_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 6
        numpad()
    def entryt2sec_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 7
        numpad()
    def entrythr_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 8
        numpad()
#     def entrythr2_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 9
#         numpad()

    t1_label = Label(settemptop_labelframe, bg='white', text='T1', fg='black', font=("Courier",20,"bold"))
    t1_label.place(x=82, y=14)
    t1oc_label = Label(settemptop_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 12,"bold"))
    t1oc_label.place(x=183, y=63)
    t1_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",38,"bold"))
    t1_entry.place(x=119,y=56)
    t1_entry.bind('<Button-1>', entryt1_click)
    t1_entry.insert(0,t1)

    t2_label = Label(settemptop_labelframe, bg='white', text='T2', fg='black', font=("Courier",20,"bold"))
    t2_label.place(x=286, y=14)
    t2oc_label = Label(settemptop_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 12,"bold"))
    t2oc_label.place(x=387, y=63)
    t2_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",38,"bold"))
    t2_entry.place(x=323,y=56)
    t2_entry.bind('<Button-1>', entryt2_click)
    t2_entry.insert(0,t2)

    t1min_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
    t1min_entry.place(x=96,y=200)
    t1min_entry.bind('<Button-1>', entryt1min_click)
    t1min_entry.insert(0,t1m)
    t1min_label = Label(settemptop_labelframe, bg='white', text='m', fg='black', font=("Courier",20,"bold"))
    t1min_label.place(x=136, y=200)
    t1sec_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
    t1sec_entry.place(x=160,y=200)
    t1sec_entry.bind('<Button-1>', entryt1sec_click)
    t1sec_entry.insert(0,t1s)
    t1sec_label = Label(settemptop_labelframe, bg='white', text='s', fg='black', font=("Courier",20,"bold"))
    t1sec_label.place(x=200, y=200)

    t2min_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
    t2min_entry.place(x=300,y=200)
    t2min_entry.bind('<Button-1>', entryt2min_click)
    t2min_entry.insert(0,t2m)
    t2min_label = Label(settemptop_labelframe, bg='white', text='m', fg='black', font=("Courier",20,"bold"))
    t2min_label.place(x=340, y=200)
    t2sec_entry = Entry(settemptop_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
    t2sec_entry.place(x=364,y=200)
    t2sec_entry.bind('<Button-1>', entryt2sec_click)
    t2sec_entry.insert(0,t2s)
    t2sec_label = Label(settemptop_labelframe, bg='white', text='s', fg='black', font=("Courier",20,"bold"))
    t2sec_label.place(x=404, y=200)

    thr_label = Label(settemptop_labelframe, bg='white', text='Threshold: ', fg='black', font=("Courier",15,"bold"))
    thr_label.place(x=145, y=260)
    thr_entry = Entry(settemptop_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
    thr_entry.place(x=277,y=255)
    thr_entry.bind('<Button-1>', entrythr_click)
    thr_entry.insert(0,thr)

#     thr2_label = Label(settemptop_labelframe, bg='white', text='Threshold 2: ', fg='black', font=("Courier",15,"bold"))
#     thr2_label.place(x=145, y=310)
#     thr2_entry = Entry(settemptop_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",20,"bold"))
#     thr2_entry.place(x=315,y=305)
#     thr2_entry.bind('<Button-1>', entrythr2_click)
#     thr2_entry.insert(0,thr2)

    def back_click():
        settemp_labelframe.place_forget()
        setfoldername()
    def thread():
        th1 = Thread(target = next_click)
        th1.start()
    def next_click():
        settemp_labelframe.place_forget()
        global thr_set, t1_set, t2_set, t1min_set, t1sec_set, t2min_set, t2sec_set
        t1_set = t1_entry.get()
        t1_set = t1_set[0:2]
        t2_set = t2_entry.get()
        t2_set = t2_set[0:2]

        t1min_set = t1min_entry.get()[0:2]
        t1sec_set = t1sec_entry.get()[0:2]
        t2min_set = t2min_entry.get()[0:2]
        t2sec_set = t2sec_entry.get()[0:2]

        thr_set = thr_entry.get()[0:4]

        global path5
        if os.path.exists(path5+"/Temperature_program.txt"):
            fc= open(path5+"/Temperature_program.txt","w")
            fc.truncate(0)
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
        else:
            fc= open(path5+"/Temperature_program.txt","w+")
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
        scanposition()
    def save_click():
        msg = messagebox.askquestion("Save Parameter", "Do you want to save?")
        if(msg=='yes'):
            fw = open("/home/pi/Spotcheck-Lamp/tempsaved.txt","w")
            fw.truncate(0)
            fw.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fw.writelines("T2="+t2_entry.get()[0:2]+"\n")

            fw.writelines("T1m="+t1min_entry.get()[0:2]+"\n")

            fw.writelines("T1s="+t1sec_entry.get()[0:2]+"\n")

            fw.writelines("T2m="+t2min_entry.get()[0:2]+"\n")

            fw.writelines("T2s="+t2sec_entry.get()[0:2]+"\n")

            fw.writelines("THR="+thr_entry.get()[0:4]+"\n")
#             if(len(thr2_entry.get())<=2):
#                 fw.writelines("THR2="+thr2_entry.get()[0:2]+".0"+"\n")
#             else:
#                 fw.writelines("THR2="+thr2_entry.get()[0:4]+"\n")


    back_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Back" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    next_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Next", height=3, width=11, borderwidth=0, command=thread)
    next_button.place(x=647,y=406)
    save_button = Button(settemp_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="Save", height=3, width=11, borderwidth=0,command=save_click)
    save_button.place(x=332,y=406)
####################################################### SET TEMPERATURES SCREEN - END ##############################################################

######################################################### SAMPLES POSITION - START #################################################################
def scanposition():
    global path0
    global path1
    global path2
    global path3
    global path4
    global path5

    global ser
    ser.flushInput()
    ser.flushOutput()
    global scanpostion_labelframe
    scanposition_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    scanposition_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(scanposition_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    scanposition_label = Label(scanposition_labelframe, bg='dodger blue', text='SAMPLES POSITION', font=("Courier",17,'bold'), width=20, height=1 )
    scanposition_label.place(x=257,y=12)

    scan_img = Image.open('/home/pi/Spotcheck-Lamp/scan.png')
    scan_width, scan_height = scan_img.size
    scale_percent = 100
    width = int(scan_width * scale_percent / 100)
    height = int(scan_height * scale_percent / 100)
    display_img = scan_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    scan_label = Label(scanposition_labelframe, bg='white',image=image_select)
    scan_label.image = image_select
    scan_label.place(x=270,y=80)

    s = ttk.Style()
    s.theme_use('clam')
    s.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
    scanposition_progressbar = ttk.Progressbar(root, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
    scanposition_progressbar.place(x=299,y=408)

    def back_click():
        try:
            try:
                camera.close()
            except Exception:
                pass
            scanposition_labelframe.place_forget()
            settemp()
        except:
            pass
    back_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Back" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)

    send_data = 'P'
    ser.write(send_data.encode())

    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()
        if(receive_data=='C'):
            global wait
            wait = 1
            scanposition_progressbar['value'] = 20
            root.update_idletasks()

    while(wait!=1):
        scanposition_progressbar['value'] = 2
        root.update_idletasks()
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 10
            root.update_idletasks()
            if(receive_data=='C'):
                scanposition_progressbar['value'] = 20
                root.update_idletasks()
                wait = 1
                break;
    while(wait==1):
        process_label = Label(scanposition_labelframe, text='Processing...', bg='white', font=("Courier",13))
        process_label.place(x=333,y=440)
        try:
            camera_capture(path4 + "/Sample_original.jpg")
        except Exception as e :
            error = messagebox.askquestion("Error: "+ str(e), "Do you want exit?", icon = "error")
            if(error=='yes'):
                root.destroy()

        image = cv2.imread(path4 + "/Sample_original.jpg")
        blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,7,21)
        gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
        thresh, binary_img = cv2.threshold(gray_img.copy(), 40, maxval=255, type=cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        print("Number of contours: " + str(len(contours)))

        contours.sort(key=lambda data:sorting_xy(data))

        contour_img = np.zeros_like(gray_img)
        bourect0 = cv2.boundingRect(contours[0])
        bourect47 = cv2.boundingRect(contours[len(contours)-1])
        global start_point
        start_point = (bourect0[0]-2, bourect0[1]-2)
        global end_point
        end_point = (bourect47[0]+bourect47[2]+2, bourect47[1]+bourect47[3]+2)
        print('Start point:', start_point)
        print('End point:', end_point)
        scanposition_progressbar['value'] = 35
        root.update_idletasks()

        global pos_result
        pos_result, pos_image = process_image(path4 + "/Sample_original.jpg")
        scanposition_progressbar['value'] = 60
        root.update_idletasks()
        sleep(1)

        output = path4 + "/Sample_processed.jpg"
        cv2.imwrite(output, pos_image)
        scanposition_progressbar['value'] = 90
        root.update_idletasks()
        sleep(1)

        scanresult_labelframe = LabelFrame(scanposition_labelframe, bg='ghost white', width=528,height = 307)
        scanresult_labelframe.place(x=248,y=60)

        label = list(range(48))
        global id_list
        def result_table(range_a, range_b, row_value):
            global samples
            j=-1
            for i in range(range_a, range_b):
                j+=1
                if(i<6):
                    t='A'+ str(i+1)
                if(i>=6 and i<12):
                    t='B'+ str(i-5)
                if(i>=12 and i<18):
                    t='C'+ str(i-11)
                if(i>=18 and i<24):
                    t='D'+ str(i-17)
                if(i>=24 and i<30):
                    t='E'+ str(i-23)
                if(i>=30 and i<36):
                    t='F'+ str(i-29)
                if(i>=36 and i<42):
                    t='G'+ str(i-35)
                if(i>=42):
                    t='H'+ str(i-41)

                if(pos_result[i]<=8):
                    label[i] = Label(scanresult_labelframe, bg='gainsboro', text=t, width=5, height=2)
                    label[i].grid(row=row_value,column=j,padx=3,pady=3)
                else:
                    label[i] = Label(scanresult_labelframe, bg='OliveDrab1', text=t, width=5, height=2)
                    label[i].grid(row=row_value,column=j,padx=3,pady=3)
                    samples += 1
        scanposition_progressbar['value'] = 100
        root.update_idletasks()

        result_table(0,6,0)
        result_table(6,12,1)
        result_table(12,18,2)
        result_table(18,24,3)
        result_table(24,30,4)
        result_table(30,36,5)
        result_table(36,42,6)
        result_table(42,48,7)
        global samples
        samplenum_label = Label(scanposition_labelframe, text='Number of Samples: ' + str(samples), fg='dodger blue', bg='white', font=("Courier",13))
        samplenum_label.place(x=293,y=432)
        scan_label.place_forget()
        scanposition_progressbar.place_forget()
        process_label.place_forget()
        wait = 0
        samples = 0
        def thread():
            th1 = Thread(target = next_click)
            th1.start()
        def next_click():
            global createclicked
            createclicked = 0
            scanposition_labelframe.place_forget()
            analysis()
        next_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Next", height=3, width=11, borderwidth=0,command=thread)
        next_button.place(x=647,y=406)
########################################################## SAMPLES POSITION - END ##################################################################

######################################################### SAMPLES ANALYSIS - START #################################################################
def analysis():
    global ser
    ser.flushInput()
    ser.flushOutput()
    global analysis_labelframe
    analysis_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    analysis_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(analysis_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    analysis_label = Label(analysis_labelframe, bg='dodger blue', text='SAMPLES ANALYSIS', font=("Courier",17,'bold'), width=20, height=1 )
    analysis_label.place(x=261,y=12)
    t_labelframe = LabelFrame(analysis_labelframe, bg='white', width=798, height=298)
    t_labelframe.place(x=0,y=70)

    t1_labelframe = LabelFrame(t_labelframe, bg='white',text="T1:"+t1_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t1_labelframe.place(x=0,y=2)
    t2_labelframe = LabelFrame(t_labelframe, bg='white',text="T2:"+t2_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t2_labelframe.place(x=199,y=2)
    t3_labelframe = LabelFrame(t_labelframe, bg='white smoke',text="T3", width=197, height=290)
    t3_labelframe.place(x=398,y=2)
    t4_labelframe = LabelFrame(t_labelframe, bg='white smoke',text="T4", width=197, height=290)
    t4_labelframe.place(x=597,y=2)
    t1wait_label = Label(t1_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t1wait_label.place(x=46,y=110)
    t2wait_label = Label(t2_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t2wait_label.place(x=46,y=110)
    temp_label = Label(analysis_labelframe, bg='white', fg='grey36', font=("Courier",20,'bold'))
    temp_label.place(x=65,y=389)

    def stop_click():
        try:
            camera.close()
        except:
            pass
        global ser
        msgbox = messagebox.askquestion('Stop the process','Do you want to stop?', icon = 'question')
        if(msgbox=='yes'):
            send_data ='S'
            ser.write(send_data.encode())
            analysis_labelframe.place_forget()
            mainscreen()

    stop_button = Button(analysis_labelframe, bg="red", font=("Courier",12,'bold'), text="STOP", height=3, width=9, borderwidth=0, command=stop_click)
    stop_button.place(x=600,y=390)
    root.update()

    send_data = "t"+ t1_set + "," + t2_set + "," + t1min_set + "," + t1sec_set + "," + t2min_set + "," + t2sec_set + "z"
    print(send_data)
    ser.write(send_data.encode())
    t0 = time.time()
    sleep(2)

    global wait
    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        if(receive_data=='Y'):
            autoprocess_label = Label(analysis_labelframe, bg='white', text="Program is processing...", fg='blue', font=("Courier",12,'bold'))
            autoprocess_label.place(x=65,y=438)
            wait = 1
    while(wait!=1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            if(receive_data=='Y'):
                autoprocess_label = Label(analysis_labelframe, bg='white', text="Program is processing...", fg='blue', font=("Courier",12,'bold'))
                autoprocess_label.place(x=65,y=438)
                wait = 1
                break

    while(wait==1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8',errors='ignore').rstrip()
            #print("Data received:", receive_data)
            if(receive_data!='C1' and receive_data!='C2' and receive_data!='C3'):
                print("Data received:", receive_data)
                temp_label['text'] = 'Temperature: '+ receive_data + chr(176)+'C'
                root.update()

            if(receive_data=='C1'):
                print("Data received:", receive_data)
                t1wait_label.place_forget()
                t1_labelframe['bg'] = atk.DEFAULT_COLOR
                t1_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t1_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t1_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Processing!', font=("Courier",9,'bold'))
                tprocess_label.place(x=59,y=112)

                global path1
                camera_capture(path1 + "/T1.jpg")

                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')

                global start_point
                global end_point
                t1_result, t1_image= process_image(path1 + "/T1.jpg")

                global path2
                output = path2 + "/T1.jpg"
                cv2.imwrite(output, t1_image)

                t1_analysis = Image.open(output)
                t1_crop = t1_analysis.crop((280-7, 81-7, 498+7, 376+7))
                crop_width, crop_height = t1_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t1_crop.resize((width,height))
                t1_display = ImageTk.PhotoImage(display_img)
                t1_label = Label(t1_labelframe, image=t1_display)
                t1_label.image = t1_display
                t1_label.place(x=0,y=1)
                root.update()

                workbook = Workbook()
                sheet = workbook.active

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"

                    sheet[pos] = t1_result[i]

                global path3
                workbook.save(path3+"/T1.xlsx")

            if(receive_data=='C2'):
                print("Data received:", receive_data)
                t2wait_label.place_forget()
                t2_labelframe['bg'] = atk.DEFAULT_COLOR
                t2_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t2_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t2_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Processing!', font=("Courier",9,'bold'))
                tprocess_label.place(x=59,y=112)

                camera_capture(path1 + "/T2.jpg")

                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                t2_result, t2_image = process_image(path1 + "/T2.jpg")
                output = path2 + "/T2.jpg"
                cv2.imwrite(output, t2_image)
                t2_analysis = Image.open(output)
                t2_crop = t2_analysis.crop((280-7, 81-7, 498+7, 376+7))
                crop_width, crop_height = t2_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t2_crop.resize((width,height))
                t2_display = ImageTk.PhotoImage(display_img)

                #t2_display = ImageTk.PhotoImage(t2_crop)
                t2_label = Label(t2_labelframe, image=t2_display)
                t2_label.image = t2_display
                t2_label.place(x=0,y=1)
                wait = 0
                root.update()
                
                stop_button.place_forget()
                temp_label.place_forget()
                autoprocess_label.place_forget()
                root.update()

                workbook = Workbook()
                sheet = workbook.active

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"

                    sheet[pos] = t2_result[i]

                workbook.save(path3+"/T2.xlsx")

                global thr_set

                workbook = Workbook()
                sheet = workbook.active

                sheet.protection.sheet = True
                sheet.protection.enable()

                #sheet["B7"].protection = Protection(locked=False, hidden=False)
                sheet["B8"].protection = Protection(locked=False, hidden=False)
                #sheet["B9"].protection = Protection(locked=False, hidden=False)

                font1 = Font(size='14', bold=True, color='00FF0000')
                font2 = Font(bold=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                img = Img('/home/pi/Spotcheck-Lamp/ps.jpg')
                img.height = 60
                img.width = 224
                img.anchor = 'B2'
                sheet.add_image(img)

                sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
                sheet["B5"] = 'LAMP TEST RESULTS TABLE'
                sheet["B5"].font = font1
                sheet.cell(row=5,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                global foldername
                sheet["B7"] = 'File name: '+ foldername
                sheet["B7"].font = font2
                sheet['B8'] = 'Operator:'
                sheet["B8"].font = font2
                global lampdir_old
                sheet['B9'] = 'Date: ' + lampdir_old[8:16]
                sheet["B9"].font = font2
                sheet['H35'] = 'Note:'
                sheet["H35"].font = font2
                sheet['H36'] = '+ N/A: None-sample position'
                sheet['H37'] = '+ N: Negative'
                sheet['J36'] = '+ P: Positive'

                for r in range(10,35):
                    for c in range(2,7):
                        sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                        sheet.cell(row=r,column=c+6).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                        sheet.cell(row=r,column=c).border = thin_border
                        sheet.cell(row=r,column=c+6).border = thin_border

                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 12
                sheet.column_dimensions['D'].width = 12
                sheet.column_dimensions['E'].width = 12
                sheet.column_dimensions['F'].width = 12
                sheet.column_dimensions['G'].width = 5
                sheet.column_dimensions['H'].width = 20
                sheet.column_dimensions['I'].width = 12
                sheet.column_dimensions['J'].width = 12
                sheet.column_dimensions['K'].width = 12
                sheet.column_dimensions['L'].width = 12
                sheet.row_dimensions[10].height = 40

                sheet['B10'] = 'PATIENT ID'
                sheet["B10"].font = font2
                sheet["B10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['C10'] = 'Spotcheck postions'
                sheet["C10"].font = font2
                sheet["C10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['D10'] = 'Spotcheck result'
                sheet["D10"].font = font2
                sheet["D10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['E10'] = 'Gel result'
                sheet["E10"].font = font2
                sheet["E10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['F10'] = 'Final result'
                sheet["F10"].font = font2
                sheet["F10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

                sheet['H10'] = 'PATIENT ID'
                sheet["H10"].font = font2
                sheet["H10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['I10'] = 'Spotcheck postions'
                sheet["I10"].font = font2
                sheet["I10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['J10'] = 'Spotcheck result'
                sheet["J10"].font = font2
                sheet["J10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['K10'] = 'Gel result'
                sheet["K10"].font = font2
                sheet["K10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['L10'] = 'Final result'
                sheet["L10"].font = font2
                sheet["L10"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

                for i in range (11,35):
                    if(i<17):
                        if(i==11):
                            sheet['C'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                            sheet['I'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')

                        sheet['C'+str(i)] = 'A'+ str(i-10)
                        sheet['I'+str(i)] = 'E'+ str(i-10)

                        sheet['B'+str(i)] = id_list[i-11]
                        sheet['H'+str(i)] = id_list[i-11+24]

                        #if(pos_result[i-11]<=15):
                        if(id_list[i-11]=='N/A'):
                            sheet['D'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11]<=float(thr_set)):
                                sheet['D'+str(i)] = 'N'
                                sheet['D'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11]>float(thr_set)):
                                sheet['D'+str(i)] = 'P'
                                sheet['D'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        #if(pos_result[i-11+24]<=15):
                        if(id_list[i-11+24]=='N/A'):
                            sheet['J'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11+24]<=float(thr_set)):
                                sheet['J'+str(i)] = 'N'
                                sheet['J'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11+24]>float(thr_set)):
                                sheet['J'+str(i)] = 'P'
                                sheet['J'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        sheet['E'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['F'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['K'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['L'+str(i)].protection = Protection(locked=False, hidden=False)

                    if(i>=17 and i<23):
                        if(i==17):
                            sheet['C'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                            sheet['I'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                        sheet['C'+str(i)] = 'B'+ str(i-16)
                        sheet['I'+str(i)] = 'F'+ str(i-16)

                        sheet['B'+str(i)] = id_list[i-11]
                        sheet['H'+str(i)] = id_list[i-11+24]

                        #if(pos_result[i-11]<=15):
                        if(id_list[i-11]=='N/A'):
                            sheet['D'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11]<=float(thr_set)):
                                sheet['D'+str(i)] = 'N'
                                sheet['D'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11]>float(thr_set)):
                                sheet['D'+str(i)] = 'P'
                                sheet['D'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        #if(pos_result[i-11+24]<=15):
                        if(id_list[i-11+24]=='N/A'):
                            sheet['J'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11+24]<=float(thr_set)):
                                sheet['J'+str(i)] = 'N'
                                sheet['J'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11+24]>float(thr_set)):
                                sheet['J'+str(i)] = 'P'
                                sheet['J'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        sheet['E'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['F'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['K'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['L'+str(i)].protection = Protection(locked=False, hidden=False)

                    if(i>=23 and i<29):
                        if(i==23):
                            sheet['C'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                            sheet['I'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                        sheet['C'+str(i)] = 'C'+ str(i-22)
                        sheet['I'+str(i)] = 'G'+ str(i-22)

                        sheet['B'+str(i)] = id_list[i-11]
                        sheet['H'+str(i)] = id_list[i-11+24]

                        #if(pos_result[i-11]<=15):
                        if(id_list[i-11]=='N/A'):
                            sheet['D'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11]<=float(thr_set)):
                                sheet['D'+str(i)] = 'N'
                                sheet['D'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11]>float(thr_set)):
                                sheet['D'+str(i)] = 'P'
                                sheet['D'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')


                        #if(pos_result[i-11+24]<=15):
                        if(id_list[i-11+24]=='N/A'):
                            sheet['J'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11+24]<=float(thr_set)):
                                sheet['J'+str(i)] = 'N'
                                sheet['J'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11+24]>float(thr_set)):
                                sheet['J'+str(i)] = 'P'
                                sheet['J'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        sheet['E'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['F'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['K'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['L'+str(i)].protection = Protection(locked=False, hidden=False)

                    if(i>=29):
                        if(i==29):
                            sheet['C'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                            sheet['I'+str(i)].fill = PatternFill(start_color='00E6B8AF', end_color='00E6B8AF', fill_type='solid')
                        sheet['C'+str(i)] = 'D'+ str(i-28)
                        sheet['I'+str(i)] = 'H'+ str(i-28)

                        sheet['B'+str(i)] = id_list[i-11]
                        sheet['H'+str(i)] = id_list[i-11+24]

                        #if(pos_result[i-11]<=15):
                        if(id_list[i-11]=='N/A'):
                            sheet['D'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11]<=float(thr_set)):
                                sheet['D'+str(i)] = 'N'
                                sheet['D'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11]>float(thr_set)):
                                sheet['D'+str(i)] = 'P'
                                sheet['D'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        #if(pos_result[i-11+24]<=15):
                        if(id_list[i-11+24]=='N/A'):
                            sheet['J'+str(i)] = 'N/A'
                        else:
                            if(t2_result[i-11+24]<=float(thr_set)):
                                sheet['J'+str(i)] = 'N'
                                sheet['J'+str(i)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                            if(t2_result[i-11+24]>float(thr_set)):
                                sheet['J'+str(i)] = 'P'
                                sheet['J'+str(i)].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

                        sheet['E'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['F'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['K'+str(i)].protection = Protection(locked=False, hidden=False)
                        sheet['L'+str(i)].protection = Protection(locked=False, hidden=False)

                sheet.print_area = 'B2:L37'
                workbook.save(path3+"/Result.xlsx")


                def thr():
                    th2 = Thread(target = viewresult_click)
                    th2.start()
                def viewresult_click():
                    viewresult_button.place_forget()
                    t1_labelframe.place_forget()
                    t2_labelframe.place_forget()
                    t3_labelframe.place_forget()
                    t_labelframe.place_forget()
                    analysis_label['text']="ANALYSIS RESULTS"

                    annotate_labelframe = LabelFrame(analysis_labelframe, bg='white', width=380, height=305)
                    annotate_labelframe.place(x=360,y=76)
                    root.update_idletasks()

                    negative_label = Label(annotate_labelframe, bg='lawn green', width=4, height=2)
                    negative_label.place(x=75,y=32)
                    negativetext_label = Label(annotate_labelframe, bg='white', text='  (N)           NEGATIVE', height=2)
                    negativetext_label.place(x=145,y=32)
                    positive_label = Label(annotate_labelframe, bg='red', width=4, height=2)
                    positive_label.place(x=75,y=82)
                    positivetext_label = Label(annotate_labelframe, bg='white', text='  (P)           POSITIVE', height=2)
                    positivetext_label.place(x=145,y=82)
                    root.update_idletasks()

                    def finish_click():
                        msgbox = messagebox.askquestion('Finish','Do you want to return to the main menu?', icon = 'question')
                        if(msgbox=='yes'):
                            global foldername
                            global covid19clicked
                            global tbclicked
                            foldername = ""
                            covid19clicked = 0
                            tb_clicked = 0
                            analysis_labelframe.place_forget()
                            mainscreen()
                    finish_button = Button(analysis_labelframe, bg="dark orange", text="FINISH", height=3, width=15, borderwidth=0, command=finish_click)
                    finish_button.place(x=480,y=396)
                    root.update_idletasks()

                    result_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 307)
                    result_labelframe.place(x=104,y=120)
                    row_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 50)
                    row_labelframe.place(x=104,y=76)
                    column_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=50,height = 307)
                    column_labelframe.place(x=62,y=120)
                    root.update_idletasks()

                    row_label = [0,0,0,0,0,0]
                    for i in range (0,6):
                        row_text = i+1
                        row_label[i] = Label(row_labelframe, text=row_text, bg='grey94', width=4, height=2)
                        row_label[i].grid(row=0,column=i,padx=2,pady=2)

                    column_label = [0,0,0,0,0,0,0,0]
                    for i in range (0,8):
                        if(i==0):
                            column_text = 'A'
                        if(i==1):
                            column_text = 'B'
                        if(i==2):
                            column_text = 'C'
                        if(i==3):
                            column_text = 'D'
                        if(i==4):
                            column_text = 'E'
                        if(i==5):
                            column_text = 'F'
                        if(i==6):
                            column_text = 'G'
                        if(i==7):
                            column_text = 'H'
                        column_label[i] = Label(column_labelframe, text=column_text, bg='grey94', width=4, height=2)
                        column_label[i].grid(row=i,column=0,padx=2,pady=2)

                    label = list(range(48))
                    def result_table(range_a,range_b, row_value):
                        j=-1
                        global pos_result
                        for i in range(range_a, range_b):
                            j+=1
                            if(id_list[i]=='N/A'):
                                label[i] = Label(result_labelframe, bg='white smoke', text='N/A', width=4, height=2)
                                label[i].grid(row=row_value,column=j,padx=2,pady=2)
                            else:
                                if(t2_result[i]<=float(thr_set)):
                                    label[i] = Label(result_labelframe, bg='lawn green', text='N', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t2_result[i]>float(thr_set)):
                                    label[i] = Label(result_labelframe, bg='red', text='P', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)

                    result_table(0,6,0)
                    result_table(6,12,1)
                    result_table(12,18,2)
                    result_table(18,24,3)
                    result_table(24,30,4)
                    result_table(30,36,5)
                    result_table(36,42,6)
                    result_table(42,48,7)
                    root.update_idletasks()

                    def detail_click():
                        if(detail_button['bg']=='lawn green'):
                            detail_button['bg']='grey94'
                            for i in range (0,48):
                                if(id_list[i]=='N/A'):
                                    label[i]['text'] = 'N/A'
                                else:
                                    if(t2_result[i]<=float(thr_set)):
                                        label[i]['text'] = 'N'
                                    if(t2_result[i]>float(thr_set)):
                                        label[i]['text'] = 'P'

                        else:
                            detail_button['bg']='lawn green'
                            for i in range (0,48):
                                if(id_list[i]=='N/A'):
                                    label[i]['text'] = 'N/A'
                                else:
                                    if(t2_result[i]<=float(thr_set)):
                                        label[i]['text'] = str('%.1f'%t2_result[i])
                                    if(t2_result[i]>float(thr_set)):
                                        label[i]['text'] = str('%.1f'%t2_result[i])
                        root.update_idletasks()
                        subprocess.call(["scrot",path3+"/result_value.jpg"])

                    detail_button = Button(analysis_labelframe, activebackground="white", bg="grey94", text="DETAIL", height=3, width=10, borderwidth=0, command=detail_click)
                    detail_button.place(x=360,y=396)
                    root.update_idletasks()
                    subprocess.call(["scrot",path3+"/result.jpg"])
#                     screenshot_img = Image.open(path3+"/result.jpg")
#                     screenshot_crop = screenshot_img.crop((60,74,352,475))
#                     screenshot_crop = screenshot_crop.save(path3+"/result.jpg")
                viewresult_button = Button(analysis_labelframe, bg="dodger blue", text="VIEW RESULT", height=3, width=15, borderwidth=0, command=thr)
                viewresult_button.place(x=327,y=394)
########################################################## SAMPLES ANALYSIS - END ##################################################################

############################################################### LOOP - START #######################################################################
while True:
    mainscreen()
    root.mainloop()
################################################################ LOOP - END ########################################################################